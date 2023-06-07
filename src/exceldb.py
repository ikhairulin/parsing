import openpyxl

import logging.config

from settings import db_path
from settings import logger_config

logging.config.dictConfig(logger_config)

console_log = logging.getLogger('console_log')
file_log = logging.getLogger('file_log')


class ExcelMaster():
    def __init__(self):
        self.db_excel_path = db_path['excel']

    def load_excel_tags_database(self) -> dict[str, tuple]:
        '''Load list of tags from excel tab and create a dict with sorted tags
        key : (category, tag_value)
        for example {'Дагот Ур': ('TES', 50),
                     'апскейл': (None, 10)}
        '''
        file_log.debug(f'Загружаю базу тэгов по адресу {self.db_excel_path}')
        wb = openpyxl.load_workbook(self.db_excel_path,
                                    read_only=True,
                                    data_only=True)

        ws = wb['j_data']
        db_tags_dict: dict = {}

        for row in ws.iter_rows(min_row=3, min_col=1, max_col=4):
            tag_list = []
            for cell in row:
                tag_list.append(cell.value)
            db_tags_dict[tag_list[1]] = (tag_list[2], tag_list[3])

        console_log.info('Загружена база тэгов длиной '
                         f'{len(db_tags_dict)} позиций')

        return db_tags_dict

    def load_twitter_authors_database(self) -> dict[str, tuple]:
        '''Load list of authors from excel tab and create a dict with nicknames
        without @-symbol
        nickname: author_name
        for example {'_pocketss': 'pocketss',
                     'arthur_vista': 'Arthur Vista'}
        '''
        file_log.debug(f'Загружаю базу тэгов по адресу {self.db_excel_path}')
        wb = openpyxl.load_workbook(self.db_excel_path,
                                    read_only=True,
                                    data_only=True)
        ws = wb['tw_data']
        db_authors_dict: dict = {}

        for row in ws.iter_rows(min_row=3, min_col=1, max_col=4):
            authors_list = []
            for cell in row:
                authors_list.append(cell.value)
            db_authors_dict[authors_list[1]] = authors_list[2]

        console_log.info('Загружена база тэгов длиной '
                         f'{len(db_authors_dict)} позиций')

        return db_authors_dict

    def load_excel_authors_database(self) -> dict[str, tuple]:
        '''
        Load list of authors from excel table and create
        a dict -> key : author_dir
        for example {'Aleriia_V': 'Lera_Pi',
                     'BRAO_art': 'BRAO'}
        '''
        file_log.debug('Загружаю базу данных художников '
                       f'по адресу {self.db_excel_path}')
        wb = openpyxl.load_workbook(self.db_excel_path,
                                    read_only=True,
                                    data_only=True)
        ws = wb['tw_data']
        db_tags_dict: dict = {}

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
            tag_list = []
            for cell in row:
                tag_list.append(cell.value)
            db_tags_dict[tag_list[1]] = tag_list[2]

        console_log.info('Загружена база тэгов длиной '
                         f'{len(db_tags_dict)} позиций')

        return db_tags_dict

    def fill_excel_database(self, authors_dict) -> None:
        """Наполняет заранее созданную excel-таблицу полученными тэгами"""
        wb = openpyxl.load_workbook(self.db_excel_path,
                                    data_only=True)
        ws = wb['tw_data']
        ws['A2']
        for i, (k, v) in enumerate(authors_dict.items()):
            ws[f"A{i+2}"] = k
            ws[f"B{i+2}"] = v
        wb.save(self.db_excel_path)
        wb.close()

        print(f"Список тэгов пополнен. Добавлено {len(authors_dict)} позиций")


def main():
    pass


if __name__ == '__main__':
    main()
